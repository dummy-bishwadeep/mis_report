import copy
import json
import re
import datetime
from scripts.utils.kairos.kairos_connection import KairosDBUtility
from scripts.logging import logger as logger
from scripts.config import DBConf


import os
import shutil
from datetime import datetime as dt

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

from scripts.core.constants import PathConstants
from scripts.core.constants.report_constants import ReportConstants, JsonConstants
from scripts.utils.kairos.kairos_connection import KairosDBUtility
from scripts.logging import logger as logger
from scripts.config import DBConf
# from scripts.core.constants import constants
import warnings
from openpyxl.drawing.image import Image
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.drawings")


class MisReport:
    def __init__(self):
        self.kairos_obj = KairosDBUtility(base_url=DBConf.KAIROS_URI, username=DBConf.KAIROS_USERNAME,
                                          password=DBConf.KAIROS_PASSWORD)


    def convert_datetime_to_epoch(self, date_time):
        """
        parameters:
            - date_time
        Returns:
            - time in epoch time format
        """
        # datetime_str = "07-31-2024 15:00:00"
        return int(datetime.datetime.strptime(date_time, "%m-%d-%Y %H:%M:%S").timestamp()) * 1000

    def convert_epoch_to_datetime(self, epoch_time):
        """
        converts epoch time format to datetime format

        Parameters:
        - epoch_time

        Returns:
        - time in datetime format
        """
        return datetime.datetime.fromtimestamp(epoch_time // 1000)

    def read_json_file(self,data):
        """
        Reads and parses a JSON file.

        Parameters:
        - file_path (str): The path to the JSON file.

        Returns:
        - dict: The JSON content as a Python dictionary.
        """
        tag_ids_list, tags_dict = list(), dict()
        try:
            if data:
                for each_key in data:
                    if each_key == "static_params":
                        for each_static_parameter in data[each_key]:
                            tag_ids_list.append(data[each_key][each_static_parameter])
                            tags_dict[each_static_parameter] = data[each_key][each_static_parameter]

                    if each_key == "lines":
                        for line_category in data[each_key]:
                            for each_line in data[each_key][line_category]:
                                for each_param in data[each_key][line_category][each_line]:
                                    for each_tag in data[each_key][line_category][each_line][each_param]:
                                        if each_tag not in ("meter_position","constant"):
                                            tag_ids_list.append(data[each_key][line_category][each_line][each_param][each_tag])
                                            tags_dict[each_line + "_" + each_tag] = data[each_key][line_category][each_line][each_param][each_tag]
                return data, tag_ids_list, tags_dict
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            return None, tag_ids_list, tags_dict

    def fetch_kairos_data(self, request_data):
        """
        This method is to get the kairos data
        :param request_data:
        :return:
        """
        logger.debug("Inside fetch_kairos_data")
        response_json = dict(
            status=False,
            message="No data within the selected date range",
            data=[]
        )
        query = {
            "metrics": [
                {
                    "tags": {
                        "c3": request_data["tag_ids_list"]
                    },
                    "name": DBConf.KAIROS_METRIC,
                    "group_by": [
                        {
                            "name": "tag",
                            "tags": [
                                "C3"
                            ]
                        }
                    ],
                    "aggregators": [
                        {
                            "name": "sum",
                            "sampling": {
                                "value": "1",
                                "unit": "days"
                            },
                            "align_sampling": True,
                            "align_start_time": True
                        }
                    ]
                }
            ],
            "plugins": [],
            "cache_time": 0,
            "time_zone": request_data["tz"],
            "start_absolute": self.convert_datetime_to_epoch(date_time=request_data["start_date"]),
            "end_absolute": self.convert_datetime_to_epoch(date_time=request_data["end_date"])
        }
        logger.info(f"Kairos Query: {query}")
        # try:
        #     kairos_response = self.kairos_obj.read(query)
        #
        #     if kairos_response["queries"][0]["results"][0]["values"]:
        #         for each_data in kairos_response["queries"][0]["results"][0]["values"]:
        #             each_data[0] = self.convert_epoch_to_datetime(epoch_time=each_data[0])
        #
        #         response_json = dict(
        #             status=True,
        #             message="Data fetched successfully",
        #             data=kairos_response["queries"][0]["results"][0]["values"]
        #         )
        # except Exception as e:
        #     logger.exception("Exception when loading the data in pareto chart")
        #     raise e
        return response_json

    def get_tag_values(self, kairos_response):
        data = kairos_response["queries"][0]["results"]
        tag_id_values = dict()
        for each_dict in data:
            tag_id = each_dict["tags"]["c3"][0]
            tag_value = each_dict["values"][0][1]
            tag_id_values[tag_id] = tag_value
        return tag_id_values

    def update_json_file(self,json_file_path, replacement_dict, output_file_path):
        """
        Read JSON file, replace keys and values based on replacement_dict, and save to a new file.
        """
        # Read the JSON file
        with open(json_file_path, 'r') as file:
            data = json.load(file)

        # Replace keys and values
        updated_data = self.replace_keys_and_values(data, replacement_dict)

        # Write the updated JSON to a new file
        with open(output_file_path, 'w') as file:
            json.dump(updated_data, file, indent=4)

    def replace_keys_and_values(self,json_data, replacement_dict):
        """
        Recursively replace keys and values in json_data based on replacement_dict.
        """
        if isinstance(json_data, dict):
            # Replace keys
            new_dict = {}
            for key, value in json_data.items():
                new_key = replacement_dict.get(key, key)
                new_value = self.replace_keys_and_values(value, replacement_dict)
                new_dict[new_key] = new_value
            return new_dict
        elif isinstance(json_data, list):
            # Replace values in lists
            return [self.replace_keys_and_values(item, replacement_dict) for item in json_data]
        elif isinstance(json_data, str):
            # Replace values in strings
            return replacement_dict.get(json_data, json_data)
        else:
            # Return other types (int, float, etc.) unchanged
            return json_data

    def evaluate_formula(self,formula, values_dict):
        """
        Evaluate the formula by replacing keys with their corresponding values from values_dict.
        """

        # Replace dictionary keys in the formula with their values
        def replace_keys_with_values(match):
            key = match.group(0)
            value = values_dict.get(key, key)
            # Evaluate expressions within the values if necessary
            try:
                # Convert fractions to float
                if '/' in value:
                    numerator, denominator = value.split('/')
                    value = float(numerator) / float(denominator)
                else:
                    value = float(value)
            except ValueError:
                raise ValueError(f"Invalid value for key '{key}': {value}")
            return str(value)

        # Regular expression to find all keys (e.g., param1, meter_position, running_hrs)
        key_pattern = re.compile(r'\b\w+\b')

        # Replace keys in the formula
        formula_with_values = key_pattern.sub(replace_keys_with_values, formula)

        # Evaluate the formula using Python's eval function
        try:
            result = eval(formula_with_values)
        except Exception as e:
            raise ValueError(f"Error evaluating formula: {e}")

        return result

    def create_final_output_json(self,edited_json):
        final_dict, category_dict = dict(), dict()
        try:
            with open(edited_json, 'r') as file:
                # Load JSON data from the file
                data = json.load(file)
            final_dict["lines"] = list()
            category_dict, formulas = dict(), dict()
            for each_key in data:
                if each_key == "formulas":
                    formulas = data["formulas"]
            color_coding = JsonConstants.color_coding

            for each_key in data:
                if each_key == "lines":
                    for line_category in data[each_key]:
                        if line_category not in category_dict:
                            category_dict[line_category] = dict()

                        for each_line in data[each_key][line_category]:
                            if each_line not in category_dict[line_category]:
                                category_dict[line_category][each_line] = dict()
                            for each_param in data[each_key][line_category][each_line]:
                                if each_param not in ('Machine'):
                                    for each_tag in data[each_key][line_category][each_line][each_param]:
                                        category_dict[line_category][each_line][each_param] = data[each_key][line_category][each_line][each_param][each_tag]
                                if each_param == 'Machine':
                                    formula = "param1*(meter_position)*(running_hrs/24) + param2 + param3*constant"
                                    formula_constants = data[each_key][line_category][each_line][each_param]
                                    result = self.evaluate_formula(formula, formula_constants)
                                    category_dict[line_category][each_line][each_param] = str(result)

                            for each_formula in formulas:
                                if each_formula not in "Machine":
                                    formula = formulas[each_formula]
                                    formula_constants = (category_dict[line_category][each_line])
                                    result = self.evaluate_formula(formula, formula_constants)
                                    category_dict[line_category][each_line][each_formula] = str(result)
            for each_line_category in category_dict:
                sample_dict = dict()
                if each_line_category in color_coding:
                    sample_dict = color_coding[each_line_category]
                sample_dict["div"] = each_line_category
                sample_dict["energy_type"] = list()
                for each_line in category_dict[each_line_category]:
                    line_dict = dict()
                    line_dict["label"] = each_line
                    line_dict["data"] = category_dict[each_line_category][each_line]
                    sample_dict["energy_type"].append(line_dict)
                final_dict["lines"].append(sample_dict)
            final_dict["static_params"] = data["static_params"]
            final_dict["static_color"] = JsonConstants.static_color
            current_ts = dt.now()
            current_ts = str(current_ts.strftime('%Y-%m-%d %H:%M:%S'))
            final_dict["metadata"] = {
    "created_by": "AYMAdmin",
    "created_on": current_ts,
    "report_date": current_ts
  }
            print("final_dict")
            print(final_dict)
            return final_dict
        except Exception as e:
            print(e)

    def generate_report(self, data=None):
        try:
            # print("data")
            # print(data)
            # data = ReportConstants.sample_data
            print(data["metadata"])
            # create output template
            metadata = data.get('metadata', {})
            created_on = metadata.get('created_on', '').replace(" ","_")
            created_on = created_on.replace(":","_")
            file_name = f'Mis_Report_{created_on}.xlsx'
            print(file_name)
            output_file_path = os.path.join(PathConstants.output_template_dir, file_name)
            shutil.copy(PathConstants.template_path, output_file_path)

            # ToDo: Add functionality for report backup count in reports directory
            # ToDo: Add proper loggers

            # Write data to excel
            self.write_data_to_excel(data=data, output_file=output_file_path)
            return file_name, output_file_path
        except Exception as generate_report_error:
            # logger.exception(f"Error while generating report: {generate_report_error}")
            raise generate_report_error

    def get_current_ts(self):
        try:
            current_ts = dt.now()
            file_ts = str(current_ts.strftime('%Y%m%d_%H%M%S'))
            current_ts = str(current_ts.strftime('%Y-%m-%d %H:%M:%S'))
            return file_ts, current_ts
        except Exception as ts_error:
            # logger.exception(f"Error while generating current timestamp: {ts_error}")
            raise ts_error

    def write_data_to_excel(self, data, output_file):
        try:
            # Load Workbook
            wb = load_workbook(output_file)
            ws = wb.active

            # Write Mis Report Metadata
            self.write_metadata(data, ws)

            # Start writing from row 13
            start_row = 13

            # Iterate over each division in the data
            for line in data['lines']:
                div = line['div']
                energy_types = line['energy_type']

                # Create fill colors for the division and energy types
                div_fill_color = PatternFill(start_color=line['div_color'].replace('#', ''), fill_type="solid")
                energy_fill_color = PatternFill(start_color=line['energy_color'].replace('#', ''), fill_type="solid")

                # Store the starting row for the division (to merge cells later)
                div_start_row = copy.deepcopy(start_row)

                # Iterate over each energy type within the division
                ws, start_row = self.write_energy_data(ws, energy_types, start_row, energy_fill_color)

                # Merge cells for the division name in column 1 and apply styles
                self.merge_div_column(ws, div_start_row, div_fill_color, div, start_row)

                # Add two extra rows as dividers between divisions
                for _ in range(2):
                    for col in range(1, 4):
                        ws.cell(row=start_row, column=col).fill = PatternFill(start_color='FFFFCC', fill_type="solid")
                    start_row += 1

            # Write the static parameters at the end of the data
            ws, start_row = self.write_static_parameters_data(data, start_row, ws)

            # Apply borders to all relevant cells
            for row in range(13, start_row):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = ReportConstants.thin_border

            # insert logo
            self.insert_aym_logo(ws)
            self.insert_kl_logo(ws)

            # Set the width of column A
            ws.column_dimensions['A'].width = 15

            # Save the modified workbook to the destination file
            wb.save(output_file)

        except Exception as report_error:
            # logger.exception(f"Error while processing data for report: {report_error}")
            raise report_error

    @staticmethod
    def write_energy_data(ws, energy_types, start_row, energy_fill_color):
        try:
            for energy in energy_types:
                # Write the energy type label in column 2 and apply styles
                ws.cell(row=start_row, column=2, value=energy['label']).font = ReportConstants.bold_font
                ws.cell(row=start_row, column=2).fill = energy_fill_color
                ws.cell(row=start_row, column=3).fill = energy_fill_color

                # Iterate over the energy data (e.g., Machine, Chiller, etc.)
                for key, value in energy['data'].items():
                    start_row += 1
                    # Write the energy data key in column 2 and apply styles
                    ws.cell(row=start_row, column=2,
                            value=key).font = ReportConstants.red_font if key == "TOTAL" else ReportConstants.bold_font
                    ws.cell(row=start_row, column=2).fill = energy_fill_color

                    # Write the corresponding value in column 3 with styles
                    cell = ws.cell(row=start_row, column=3, value=value)
                    cell.font = ReportConstants.red_font if key == "TOTAL" else ReportConstants.normal_font
                    cell.alignment = ReportConstants.center_align
                    cell.fill = PatternFill(start_color='FFFFFF',
                                            fill_type='solid') if key == "TOTAL" else energy_fill_color

                # Move to the next row for the next energy type
                start_row += 1
            return ws, start_row
        except Exception as energy_error:
            raise energy_error

    @staticmethod
    def write_static_parameters_data(data, start_row, ws):
        try:
            # Write the static parameters
            static_fill_color = PatternFill(start_color=data['static_color'].replace('#', ''), fill_type="solid")
            for key, value in data['static_params'].items():
                # Write the parameter name in column 2 and the value in column 3
                ws.cell(row=start_row, column=2, value=key).font = ReportConstants.bold_font
                ws.cell(row=start_row, column=3).value = value
                ws.cell(row=start_row, column=3).font = ReportConstants.normal_font
                for col in range(1, 4):
                    ws.cell(row=start_row, column=col).fill = static_fill_color
                start_row += 1
            return ws, start_row
        except Exception as xlsx_error:
            # logger.exception(f"Error while writing static parameters data: {xlsx_error}")
            raise xlsx_error

    @staticmethod
    def merge_div_column(ws, div_start_row, div_fill_color, div, start_row):
        try:
            ws.merge_cells(start_row=div_start_row, start_column=1, end_row=start_row - 1, end_column=1)
            div_cell = ws.cell(row=div_start_row, column=1, value=div)
            div_cell.border = ReportConstants.thin_border
            div_cell.alignment = ReportConstants.rotate_text
            div_cell.font = ReportConstants.div_font
            div_cell.fill = div_fill_color
        except Exception as merge_error:
            raise merge_error

    @staticmethod
    def write_metadata(data, ws):
        try:
            metadata = data.get('metadata', {})
            created_by = metadata.get('created_by', '')
            created_on = metadata.get('created_on', '')
            report_date = metadata.get('report_date', '')
            metadata = [created_by, created_on, report_date]

            start_row = 8
            for index in range(0, 3):
                ws.cell(row=start_row, column=3).value = metadata[index]
                ws.cell(row=start_row, column=3).alignment = ReportConstants.center_align
                start_row += 1

        except Exception as metadata_error:
            raise metadata_error

    @staticmethod
    def insert_kl_logo(ws):
        try:
            img_path = Path(PathConstants.kl_logo)
            img = Image(img_path)
            img.anchor = 'C1'

            # Add the image to the worksheet
            ws.add_image(img)
        except Exception as logo_error:
            raise logo_error

    @staticmethod
    def insert_aym_logo(ws):
        try:
            img_path = Path(PathConstants.aym_logo)
            img = Image(img_path)
            img.anchor = 'A2'

            # Add the image to the worksheet
            ws.add_image(img)
        except Exception as logo_error:
            raise logo_error
